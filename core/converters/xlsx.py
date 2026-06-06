from pathlib import Path
import openpyxl

def _rows_to_gfm(rows: list[list]) -> str:
    if not rows:
        return ""
    header = "| " + " | ".join(str(c) for c in rows[0]) + " |"
    separator = "| " + " | ".join(["---"] * len(rows[0])) + " |"
    body_lines = ["| " + " | ".join(str(c) for c in row) + " |" for row in rows[1:]]
    return "\n".join([header, separator] + body_lines)

def convert_xlsx(source: Path, assets_dir: Path) -> tuple[str, list[Path]]:
    wb = openpyxl.load_workbook(str(source), data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([cell if cell is not None else "" for cell in row])
        if not rows:
            continue
        table_md = _rows_to_gfm(rows)
        sections.append(f"## {sheet_name}\n\n{table_md}")

    return "\n\n".join(sections), []
